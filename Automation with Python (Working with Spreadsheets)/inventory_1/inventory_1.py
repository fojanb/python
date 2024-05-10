# A python library to read/write excel xlsx/xlsm file.
import openpyxl
import colors
# Let's load the inventory_1.xlsx file:
inventory_1 = openpyxl.load_workbook("inventory_1.xlsx")
# workbook is the whole .xlsx file and worksheet is each sheet inside the .xlsx file i.e. Sheet1
# Imagine the excel file as a dictionary -> inventory_1.xlsx = {"Sheet1":[]}
# Let's read and print the Sheet1:
sheet_1 = inventory_1["Sheet1"]
# print(sheet_1)
# How many rows in the Sheet1 ?
# print(sheet_1.max_row)
# How many columns in the Sheet1 ?
# print(sheet_1.max_column)
# ✿------------------- Excercise 1 -------------------✿
# List each company with respective product count
product_per_supplier = {}
for row in range(2,sheet_1.max_row + 1):
    supplier = sheet_1.cell(row,4).value
    if(supplier in product_per_supplier):
        product_per_supplier[supplier] = product_per_supplier[supplier] + 1
    else:
        product_per_supplier[supplier] = 1
print(colors.RED +"~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n" + colors.RESET)
print(colors.MAGENTA + f"✿ List each company with respective product count : {product_per_supplier}\n" + colors.RESET)
print(colors.CYAN + f"✿ Supplier counts: {len(product_per_supplier)}\n" + colors.RESET)

# ✿------------------- Excercise 2 -------------------✿
# List products with inventory less than 10
inventory_under_10 = {}
for row in range(2,sheet_1.max_row + 1):
    if(int(sheet_1.cell(row,2).value) < 10):
        key_name = "product_" + str(sheet_1.cell(row,1).value)
        inventory_under_10[key_name] = sheet_1.cell(row,2).value
    
print(colors.YELLOW + f"✿ List products with inventory less than 10 : {inventory_under_10}\n" + colors.RESET)
# ✿------------------- Excercise 3 -------------------✿
# List each company with respective total inventory value
total_inventory_value_per_supplier = {}
for row in range(2,sheet_1.max_row + 1):
    supplier = sheet_1.cell(row,4).value
    current_value = sheet_1.cell(row,2).value * sheet_1.cell(row,3).value
    if supplier in total_inventory_value_per_supplier:
        total_inventory_value_per_supplier[supplier] = total_inventory_value_per_supplier[supplier] + current_value
    else:
        total_inventory_value_per_supplier[supplier] = current_value
    
print(colors.GREEN + f"✿ List each company with respective total inventory value : {total_inventory_value_per_supplier}\n" + colors.RESET)

# ✿------------------- Excercise 4 -------------------✿
# Write to spreadsheet : calculate and write inventory value for each product into spreadsheet


for row in range(2,sheet_1.max_row + 1):
    total_inventory_price = sheet_1.cell(row,5)
    supplier = sheet_1.cell(row,4).value
    current_value = sheet_1.cell(row,2).value * sheet_1.cell(row,3).value
    total_inventory_price.value  = current_value

    if supplier in total_inventory_value_per_supplier:
        total_inventory_value_per_supplier[supplier] = total_inventory_value_per_supplier[supplier] + current_value
    else:
        total_inventory_value_per_supplier[supplier] = current_value
    
inventory_1.save("inventory_1_with_total_inventory_price.xlsx")
print(colors.RED +"~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~" + colors.RESET)
