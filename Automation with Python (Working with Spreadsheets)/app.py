import openpyxl
from colors import CYAN,MAGENTA,BRIGHT_YELLOW,RED,BRIGHT_RED,RESET
# Read an excel file from the current directory on your local machine.
# sample_data_all = SampleData.xlsx with all of the sheets inside of it.
sample_data_all = openpyxl.load_workbook("SampleData.xlsx")
# Select the desired 'sheet' we want to work on it.
# SampleData.xlsx = {"Instructions":[], "SalesOrders":[] , "MyLinks":[]}.
SalesOrders = sample_data_all["SalesOrders"]
purchaseCount_totalExpense_per_buyer = {}
buyers = []
def get_purchaseCount_totalExpense_per_buyer():
    for row in range(SalesOrders.max_row - 1) :
        buyer = SalesOrders.cell(row + 2,3).value;
        buyer_total_expenses = SalesOrders.cell(row + 2,7).value;
        buyers.append(buyer)
        if buyer in purchaseCount_totalExpense_per_buyer:
            purchaseCount_totalExpense_per_buyer[buyer]["purchases_count"]= purchaseCount_totalExpense_per_buyer[buyer]["purchases_count"] + 1
            purchaseCount_totalExpense_per_buyer[buyer]["total_expenses"] = purchaseCount_totalExpense_per_buyer[buyer]["total_expenses"] + buyer_total_expenses
        else:
            # Adding new buyer to purchaseCount_totalExpense_per_buyer
            purchaseCount_totalExpense_per_buyer[buyer] = {}
            purchaseCount_totalExpense_per_buyer[buyer]["purchases_count"]= 1
            purchaseCount_totalExpense_per_buyer[buyer]["total_expenses"] = buyer_total_expenses
get_purchaseCount_totalExpense_per_buyer()
# Possible guerys:
print(MAGENTA + "1- Total number of purchases per buyer." 
      + BRIGHT_YELLOW + "\n✿ Answer: "+ CYAN + f"{purchaseCount_totalExpense_per_buyer}" + RESET)
print(MAGENTA + "\n2- Total number of purchases per buyer sorted alphabetically." 
      + BRIGHT_YELLOW + "\n✿ Answer: "+ CYAN + f"{dict(sorted(purchaseCount_totalExpense_per_buyer.items()))}" + RESET)
print(MAGENTA + "\n3- How many buyers in the excel sheet?" 
      + BRIGHT_YELLOW + "\n✿ Answer: " + CYAN + f"{len(purchaseCount_totalExpense_per_buyer)}" + RESET)
print(MAGENTA + "\n4- Sort the buyers name alphabetically." 
      + BRIGHT_YELLOW + "\n✿ Answer: "+ CYAN + f"{sorted(list(set(buyers)))}" + RESET)


# ------ ✿ Functions ✿ -------

# Find the minimum number of total purchases and return bueyrs name who made this minimum amount of purchases :
def find_minimum_purchase():
    minimum_purchase = min(purchaseCount_totalExpense_per_buyer.values())
    buyers_with_minimum_purchases = {}
    for key in purchaseCount_totalExpense_per_buyer:
        if purchaseCount_totalExpense_per_buyer[key]["purchases_count"] == minimum_purchase:
            buyers_with_minimum_purchases[key] = minimum_purchase
    print(BRIGHT_RED +"Minimum paurchases is "+ f"{minimum_purchase} 👉 " 
          + f"{buyers_with_minimum_purchases}" + RESET)
    
# -----------------------------
    
# Get a buyer's name from user and find/return the number of total purchases that he/she made :
def find_number_of_purchases_by_buyer_name():
    a_buyer_name = input("Please enter the buyer's name to see how many purchases he/she made so far in total:")
    if a_buyer_name in purchaseCount_totalExpense_per_buyer:
        total_purchases = purchaseCount_totalExpense_per_buyer[a_buyer_name]
        print(BRIGHT_YELLOW + f"{a_buyer_name} has already made {total_purchases} purchases in total." + RESET)
    else:
        print(RED + "Not existed buyer." + RESET)
# for loop is for iterating over a list.range() will create a list for us and  then
# we can use the for loop

# -----------------------------

# How much money each buyer spent in total?
